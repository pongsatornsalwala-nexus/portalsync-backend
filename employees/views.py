from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from django.db.models import Count, Q
from .models import Employee
from .serializers import EmployeeSerializer, EmployeeListSerializer

class EmployeeViewSet(viewsets.ModelViewSet):
    """
    API endpoint for Employees

    Provides CRUD operations plus custom endpoints for dashboard stats.
    """
    def get_queryset(self):
        queryset = Employee.objects.select_related('worksite').all()

        reg_type = self.request.query_params.get('registration_type')
        month = self.request.query_params.get('month')
        worksite_id = self.request.query_params.get('worksite')
        benefit = self.request.query_params.get('benefit')

        if reg_type:
            queryset = queryset.filter(registration_type=reg_type)

        if month:
            year, mon = month.split('-')
            queryset = queryset.filter(
                created_at__year=year,
                created_at__month=mon
            )

        if worksite_id:
            queryset = queryset.filter(worksite_id=worksite_id)

        if benefit == 'SSF':
            queryset = queryset.filter(has_ssf=True)
        elif benefit == 'AIA':
            queryset = queryset.filter(has_aia=True)

        return queryset

    def get_serializer_class(self):
        """Use simplified serializer for list view, full serializer for detail view."""
        if self.action == 'list':
            return EmployeeListSerializer
        return EmployeeSerializer

    @action(detail = False, methods = ['get'])
    def count(self, request):
        """
        GET /api/employees/count/
        Returns total count of employees.
        """
        count = self.get_queryset().count()
        return Response({'count': count})
    
    @action(detail = False, methods = ['get'])
    def stats(self, request):
        """
        GET /api/employees/stats/
        Returns dashboard statistics:
        - Total employees
        - New joiners (this month)
        - Resignations (this month)
        - Pending actions
        - SSF queue counts
        - AIA queue counts
        """
        from django.utils.timezone import now
        from datetime import timedelta

        # Get date range for "this month"
        today = now()
        month_start = today.replace(day = 1)

        total = self.get_queryset().count()

        # New joiners this month
        new_joiners = self.get_queryset().filter(
            employment_date__gte = month_start,
            registration_type = 'REGISTER_IN'
        ).count()

        # Resignations this month
        resignations = self.get_queryset().filter(
            updated_at__gte = month_start,
            registration_type = 'REGISTER_OUT'
        ).count()

        # Pending actions (not yet verified)
        pending = (
            self.get_queryset().filter(has_ssf=True, ssf_activated=False, is_exiting_ssf=False).count() +
            self.get_queryset().filter(has_aia=True, aia_activated=False, is_exiting_aia=False).count() +
            self.get_queryset().filter(is_exiting_ssf=True, ssf_exit_status__in=['IMPORTED', 'PENDING']).count() +
            self.get_queryset().filter(is_exiting_aia=True, aia_exit_status__in=['IMPORTED', 'PENDING']).count()
        )

        # SSF queue counts
        ssf_register_in = self.get_queryset().filter(
            has_ssf = True,
            ssf_activated = False,
            is_exiting_ssf = False,
        ).count()

        ssf_register_out = self.get_queryset().filter(
            is_exiting_ssf = True,
            ssf_exit_status__in = ['IMPORTED', 'PENDING']
        ).count()

        # AIA queue counts
        aia_register_in = self.get_queryset().filter(
            has_aia = True,
            aia_activated = False,
            is_exiting_aia = False,
        ).count()

        aia_register_out = self.get_queryset().filter(
            is_exiting_aia = True,
            aia_exit_status__in = ['IMPORTED', 'PENDING']
        ).count()

        return Response({
            'total_employees': total,
            'new_joiners': new_joiners,
            'resignations': resignations,
            'pending_actions': pending,
            'ssf_queue': {
                'register_in': ssf_register_in,
                'register_out': ssf_register_out,
            },
            'aia_queue': {
                'register_in': aia_register_in,
                'register_out': aia_register_out,
            }
        })
    
    @action(detail = False, methods = ['get'])
    def by_worksite(self, request):
        """
        GET /api/employees/by_worksite/?worksite_id = 1
        Returns employees filtered by worksite.
        """
        worksite_id = request.query_params.get('worksite_id')
        if worksite_id:
            employees = self.get_queryset().filter(worksite_id = worksite_id)
        else:
            employees = self.get_queryset().all()

        serializer = self.get_serializer(employees, many = True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'])
    def download_template(self, request):
        """
        GET /api/employees/download_template/
        Downloads the Excel template for batch upload
        """
        import os
        from django.http import FileResponse
        from django.conf import settings

        # Build the path to the template file
        template_path = os.path.join(
            settings.BASE_DIR,
            'templates',
            'SSF_Entry.xlsx'
        )

        if not os.path.exists(template_path):
            return Response(
                {'error': 'Template file not found.'},
                status = status.HTTP_404_NOT_FOUND
            )

        # Open and return the file
        file = open(template_path, 'rb')
        response = FileResponse(
            file,
            content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment;  filename="SSF_Entry.xlsx"'

        return response

    @action(detail=False, methods=['get'])
    def download_aia_template(self, request):
        import os
        from django.http import FileResponse
        from django.conf import settings

        template_path = os.path.join(
            settings.BASE_DIR,
            'templates',
            'AIA_Entry.xlsx'
        )

        if not os.path.exists(template_path):
            return Response(
                {'error': 'Template file not found.'},
                status=status.HTTP_404_NOT_FOUND
            )

        file = open(template_path, 'rb')
        response = FileResponse(
            file,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="AIA_Entry_Template.xlsx"'
        return response

    @action(detail=False, methods=['post'])
    def bulk_upload(self, request):
        """
        POST /api/employees/bulk_upload/
        Uploads and processes Excel file to create employees in bulk
        """
        import openpyxl
        from datetime import datetime

        # Get the uploaded file
        uploaded_file = request.FILES.get('file')
        if not uploaded_file:
            return Response(
                {'error': 'No file uploaded'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Get context data
        worksite_id = request.data.get('worksite_id')
        registration_type = request.data.get('registration_type')
        benefit_type = request.data.get('benefit_type')

        if not worksite_id:
            return Response(
                {'error': 'Worksite ID is required'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            # Load the Excel file
            wb = openpyxl.load_workbook(uploaded_file)
            ws = wb.active

            created_count = 0
            errors = []

            # Parse date: "25/6/2547" -> convert Buddhist year to Gregorian
            def parse_buddhist_date(date_str):
                if not date_str:
                    return None
                try:
                    # Handle different formats
                    if isinstance(date_str, str):
                        parts = date_str.split('/')
                        if len(parts) == 3:
                            day = int(parts[0])
                            month = int(parts[1])
                            year = int(parts[2]) - 543 # Convert Buddhist to Gregorian
                            return f'{year:04d}-{month:02d}-{day:02d}'
                    return None
                except:
                    return None

            # Start reading from row 5 (skip header rows 1-4)
            for row_num in range(5, ws.max_row + 1):
                try:
                    # Read data from each column
                    prefix_thai = ws.cell(row=row_num, column=2).value # B - คำนำหน้า
                    first_name = ws.cell(row=row_num, column=3).value # C - ชื่อ
                    last_name = ws.cell(row=row_num, column=4).value # D - นามสกุล
                    ssf_status_raw = ws.cell(row=row_num, column=5).value # E - SSF แจ้งเข้า
                    national_id = ws.cell(row=row_num, column=6).value # F - เลขประจำตัวประชาชน
                    ssf_type_raw = ws.cell(row=row_num, column=7).value # G - ประเภท
                    hospital1 = ws.cell(row=row_num, column=8).value # H - โรงพยาบาล 1
                    hospital2 = ws.cell(row=row_num, column=9).value # I - โรงพยาบาล 2
                    hospital3 = ws.cell(row=row_num, column=10).value # J - โรงพยาบาล 3
                    dob = ws.cell(row=row_num, column=11).value # K - วันเกิด
                    employment_date = ws.cell(row=row_num, column=12).value # L - วันเริ่มงาน
                    aia_status_raw = ws.cell(row=row_num, column=13).value # M - AIA

                    # Skip empty rows
                    if not first_name or not national_id:
                        continue

                    # Map Excel status values to Django status values
                    status_map = {
                        'Imported': 'IMPORTED',
                        'Pending': 'PENDING',
                        'Registered': 'REGISTERED'
                    }
                    ssf_status = status_map.get(str(ssf_status_raw).strip()) if ssf_status_raw else 'IMPORTED'
                    aia_status = status_map.get(str(aia_status_raw).strip()) if aia_status_raw else 'IMPORTED'

                    # Map Thai prefix to English
                    prefix_map = {
                        'นาย': 'mr',
                        'นาง': 'mrs',
                        'นางสาว': 'ms',
                    }
                    prefix = prefix_map.get(str(prefix_thai).strip()) if prefix_thai else None
                    
                    # Determine gender from prefix
                    gender = 'male' if prefix == 'mr' else 'female'

                    # Parse ssf_type - strip whitespace just in case
                    ssf_type = str(ssf_type_raw).strip() if ssf_type_raw else '1-03'

                    # 1-03/1 means no hospital selection needed - clear them
                    if ssf_type == '1-03/1':
                        hospital1 = hospital2 = hospital3 = None

                    # Create employee
                    employee_data = {
                        'id_card': str(national_id).replace('-', ''),
                        'prefix': prefix,
                        'first_name': str(first_name).strip(),
                        'last_name': str(last_name).strip() if last_name else '',
                        'date_of_birth': parse_buddhist_date(dob),
                        'gender': gender,
                        'nationality': 'thai',
                        'employment_date': parse_buddhist_date(employment_date) or datetime.now().date(),
                        'worksite_id': int(worksite_id),
                        'has_ssf': benefit_type == 'SSF',
                        'has_aia': benefit_type == 'AIA',
                        'registration_type': registration_type,
                        'hospital_choice_1': hospital1,
                        'hospital_choice_2': hospital2,
                        'hospital_choice_3': hospital3,
                        'ssf_type': ssf_type,
                        'ssf_status': ssf_status if benefit_type == 'SSF' else None,
                        'aia_status': aia_status if benefit_type == 'AIA' else None,
                    }

                    # Create the employee
                    Employee.objects.create(**employee_data)
                    created_count += 1

                except Exception as e:
                    errors.append(f'Row {row_num}: {str(e)}')
                    continue

            return Response({
                'success': True,
                'created_count': created_count,
                'errors': errors if errors else None,
            })

        except Exception as e:
            return Response(
                {'error': f'Failed to process file: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=False, methods=['post'])
    def aia_entry_upload(self, request):
        """
        POST /api/employees/aia_entry_upload/
        Reads an official AIA Entry (Addition) xlsx and creates or updates employees.
        Datarows start at row 21. Dates are already A.D. - no Buddhist calendar conversion.
        """
        import openpyxl

        uploaded_file = request.FILES.get('file')
        if not uploaded_file:
            return Response({'error': 'No file uploaded'}, status=status.HTTP_400_BAD_REQUEST)

        worksite_id = request.data.get('worksite_id')
        if not worksite_id:
            return Response({'error': 'Worksite ID is required'}, status=status.HTTP_400_BAD_REQUEST)

        # Helper: reassemble a date from 3 separate D/M/Y cells (all A.D.)
        def assemble_date(daty, month, year):
            try:
                if not all([day, month, year]):
                    return None
                return f'{int(year):04d}-{int(month):02d}-{int(day):02d}'
            except:
                return None

        # Helper: map AIA's Sex codes to the model's values
        def parse_gender(val):
            return 'male' if str(val).strip().upper() == 'M' else 'female'

        # Helper: map AIA's Marital Status codes
        def parse_marital(val):
            mapping = {'S': 'single', 'M': 'married', 'D': 'divorced'}
            return mapping.get(str(val).strip().upper())

        try:
            wb = openpyxl.load_workbook(uploaded_file)
            ws = wb.active

            created_count = 0
            updated_count = 0
            errors = []

            # Row 21 onwards is where real employee data lives
            for row_num in range(21, ws.max_row + 1):
                try:
                    id_card = ws.cell(row=row_num, column=11).value
                    first_name = ws.cell(row=row_num, column=4).value

                    # Skip blank rows - AIA froms always have lots of empty trailing
                    if not first_name and not id_card:
                        continue

                    id_card = str(id_card).strip().replace('-', '') if id_card else None
                    if not id_card:
                        errors.append(f'Row {row_num}: Missing ID card number')
                        continue

                    employee_data = {
                        'employee_no': str(ws.cell(row=row_num, column=1).value or '').strip(),
                        'first_name': str(first_name).strip() if first_name else '',
                        'last_name': str(ws.cell(row=row_num, column=5).value or '').strip(),
                        'gender': parse_gender(ws.cell(row=row_num, column=6).value or ''),
                        'marital_status': parse_marital(ws.cell(row=row_num, column=7).value),
                        'date_of_birth': assemble_date(
                            ws.cell(row=row_num, column=8).value,
                            ws.cell(row=row_num, column=9).value,
                            ws.cell(row=row_num, column=10).value,
                        ),
                        'nationality': str(ws.cell(row=row_num, column=12).value or '').strip(),
                        'employment_date': assemble_date(
                            ws.cell(row=row_num, column=13).value,
                            ws.cell(row=row_num, column=14).value,
                            ws.cell(row=row_num, column=15).value,
                        ),
                        'salary': ws.cell(row=row_num, column=16).value,
                        'designation': str(ws.cell(row=row_num, column=17).value or '').strip(),
                        'plan': str(ws.cell(row=row_num, column=18).value or '').strip(),
                        'bank_name': str(ws.cell(row=row_num, column=19).value or '').strip(),
                        'bank_account': str(ws.cell(row=row_num, column=20).value or '').strip(),
                        'effective_date': assemble_date(
                            ws.cell(row=row_num, column=25).value,
                            ws.cell(row=row_num, column=26).value,
                            ws.cell(row=row_num, column=27).value,
                        ),
                        # AIA Entry is always an addition (inbound) for AIA benefit
                        'has_aia': True,
                        'aia_status': 'IMPORTED',
                        'registration_type': 'REGISTER_IN',
                        'worksite_id': int(worksite_id),
                    }

                    # update_or_create: looks up by id_card, applies employee_data as defaults.
                    # If found -> updates. If not found -> creates. Returns (object, created_bool).
                    _, created = Employee.objects.update_or_create(
                        id_card=id_card,
                        defaults=employee_data,
                    )

                    if created:
                        created_count += 1
                    else:
                        updated_count += 1
                    
                except Exception as e:
                    errors.append(f'Row {row_num}: {str(e)}')
                    continue

            return Response({
                'success': True,
                'created_count': created_count,
                'updated_count': updated_count,
                'errors': errors if errors else None,
            })

        except Exception as e:
            return Response(
                {'error', f'Failed to process file: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=True, methods=['patch'])
    def upload_document(self, request, pk=None):
        """
        PATCH /api/employees/{id}/upload_document/
        Upload a document file for an employee
        Expects: file (the file), file_type ('national_id', 'bank_book', 'ceb_form')
        """
        employee = self.get_object()

        # Get the uploaded file
        uploaded_file = request.FILES.get('file')
        file_type = request.data.get('file_type')

        if not uploaded_file:
            return Response(
                {'error': 'No file uploaded'},
                status=status.HTTP_400_BAD_REQUEST
            )

        if not file_type:
            return Response(
                {'error': 'file_type parameter is required'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Map file_type to model field
        field_mapping = {
            'national_id': 'national_id_file',
            'bank_book': 'bank_book_file',
            'ceb_form': 'ceb_form_file',
        }

        field_name = field_mapping.get(file_type)
        if not field_name:
            return Response(
                {'error': f'Invalid file_type: {file_type}'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Delete old file if it exists
        old_file = getattr(employee, field_name)
        if old_file:
            old_file.delete(save=False)
        
        # Save the new file
        setattr(employee, field_name, uploaded_file)
        employee.save()

        # Return updated employee data
        serializer = self.get_serializer(employee)
        return Response(serializer.data)